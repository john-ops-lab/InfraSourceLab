"""导出构建器。

- JSON：元数据、CI 和关系；
- CSV：ZIP 打包的多文件（summary、spec、CI、关系、质量报告）；
- XLSX：摘要、各 CI 类型、关系和质量报告工作表。

安全约束：
- 对以 =、+、-、@ 开头的单元格做公式注入转义；
- 导出不包含 API Key、AI Key 或内部 search_text；
- 文件名由数据集 ID 生成。
"""

import csv
import io
import json
import zipfile

from openpyxl import Workbook
from sqlalchemy.orm import Session

from ..datasets.service import dataset_summary, iter_all_cis, iter_all_relations
from ..db.models import Dataset

_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _safe_cell(value) -> str:
    text = "" if value is None else str(value)
    if text.startswith(_FORMULA_PREFIXES):
        return "'" + text
    return text


def _ci_payload(record) -> dict:
    return {
        "id": record.ci_id,
        "type": record.type,
        "name": record.name,
        "attributes": json.loads(record.attributes_json),
        "tags": json.loads(record.tags_json),
    }


def _relation_payload(record) -> dict:
    return {
        "id": record.relation_id,
        "type": record.type,
        "from_id": record.from_ci_id,
        "to_id": record.to_ci_id,
        "attributes": json.loads(record.attributes_json),
    }


def _iter_quality_rows(quality_report: list[dict]):
    for report in quality_report:
        sources = report.get("source_by_duplicate_id", {})
        applied_value = report.get("applied_value")
        affected_ids = report.get("affected_ids", [])
        if not affected_ids:
            yield [
                report.get("kind", ""),
                report.get("ci_type", ""),
                report.get("field") or "",
                report.get("requested_count", 0),
                report.get("affected_count", 0),
                "",
                "",
                applied_value if applied_value is not None else "",
            ]
            continue
        for ci_id in affected_ids:
            yield [
                report.get("kind", ""),
                report.get("ci_type", ""),
                report.get("field") or "",
                report.get("requested_count", 0),
                report.get("affected_count", 0),
                ci_id,
                sources.get(ci_id, ""),
                applied_value if applied_value is not None else "",
            ]


def build_json_export(session: Session, dataset: Dataset) -> tuple[bytes, str]:
    summary = dataset_summary(session, dataset)
    summary["quality_report"] = json.loads(dataset.quality_report_json)
    payload = {
        "dataset": summary,
        "cis": [_ci_payload(ci) for ci in iter_all_cis(session, dataset.id)],
        "relations": [_relation_payload(rel) for rel in iter_all_relations(session, dataset.id)],
    }
    body = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    return body, f"dataset-{dataset.id}.json"


def build_csv_export(session: Session, dataset: Dataset) -> tuple[bytes, str]:
    """CSV 导出为 ZIP：summary.csv + 每个类型一个 ci_<type>.csv + relations.csv。"""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        summary = dataset_summary(session, dataset)
        summary_rows = io.StringIO()
        writer = csv.writer(summary_rows)
        writer.writerow(["key", "value"])
        writer.writerow(["id", dataset.id])
        writer.writerow(["name", _safe_cell(dataset.name)])
        writer.writerow(["seed", dataset.seed])
        writer.writerow(["generator_version", dataset.generator_version])
        writer.writerow(["record_count", dataset.record_count])
        writer.writerow(["relation_count", dataset.relation_count])
        writer.writerow([])
        writer.writerow(["ci_type", "count"])
        for ci_type, count in summary["ci_counts_by_type"].items():
            writer.writerow([ci_type, count])
        archive.writestr("summary.csv", summary_rows.getvalue().encode("utf-8"))
        archive.writestr(
            "spec.json",
            json.dumps(summary["spec"], ensure_ascii=False, indent=2).encode("utf-8"),
        )

        quality_rows = io.StringIO()
        writer = csv.writer(quality_rows)
        writer.writerow(
            [
                "kind",
                "ci_type",
                "field",
                "requested_count",
                "affected_count",
                "ci_id",
                "source_id",
                "applied_value",
            ]
        )
        for row in _iter_quality_rows(json.loads(dataset.quality_report_json)):
            writer.writerow([_safe_cell(value) for value in row])
        archive.writestr("quality_report.csv", quality_rows.getvalue().encode("utf-8"))

        by_type: dict[str, list] = {}
        for record in iter_all_cis(session, dataset.id):
            by_type.setdefault(record.type, []).append(record)

        for ci_type, records in by_type.items():
            attribute_keys: list[str] = []
            seen_keys: set[str] = set()
            parsed = []
            for record in records:
                attributes = json.loads(record.attributes_json)
                parsed.append((record, attributes))
                for key in attributes:
                    if key not in seen_keys:
                        seen_keys.add(key)
                        attribute_keys.append(key)

            rows = io.StringIO()
            writer = csv.writer(rows)
            writer.writerow(["id", "name", *attribute_keys, "tags"])
            for record, attributes in parsed:
                writer.writerow(
                    [
                        record.ci_id,
                        _safe_cell(record.name),
                        *[_safe_cell(attributes.get(key)) for key in attribute_keys],
                        json.dumps(json.loads(record.tags_json), ensure_ascii=False),
                    ]
                )
            archive.writestr(f"ci_{ci_type}.csv", rows.getvalue().encode("utf-8"))

        relation_rows = io.StringIO()
        writer = csv.writer(relation_rows)
        writer.writerow(["id", "type", "from_id", "to_id", "attributes"])
        for record in iter_all_relations(session, dataset.id):
            writer.writerow(
                [
                    record.relation_id,
                    record.type,
                    record.from_ci_id,
                    record.to_ci_id,
                    record.attributes_json,
                ]
            )
        archive.writestr("relations.csv", relation_rows.getvalue().encode("utf-8"))

    return buffer.getvalue(), f"dataset-{dataset.id}.zip"


def build_xlsx_export(session: Session, dataset: Dataset) -> tuple[bytes, str]:
    summary = dataset_summary(session, dataset)
    workbook = Workbook(write_only=False)

    overview = workbook.active
    overview.title = "摘要"
    overview.append(["key", "value"])
    overview.append(["id", dataset.id])
    overview.append(["name", _safe_cell(dataset.name)])
    overview.append(["seed", dataset.seed])
    overview.append(["generator_version", dataset.generator_version])
    overview.append(["record_count", dataset.record_count])
    overview.append(["relation_count", dataset.relation_count])
    overview.append([])
    overview.append(["ci_type", "count"])
    for ci_type, count in summary["ci_counts_by_type"].items():
        overview.append([ci_type, count])

    by_type: dict[str, list] = {}
    for record in iter_all_cis(session, dataset.id):
        by_type.setdefault(record.type, []).append(record)
    for ci_type, records in by_type.items():
        sheet = workbook.create_sheet(f"CI_{ci_type}"[:31])
        attribute_keys: list[str] = []
        seen: set[str] = set()
        parsed = []
        for record in records:
            attributes = json.loads(record.attributes_json)
            parsed.append((record, attributes))
            for key in attributes:
                if key not in seen:
                    seen.add(key)
                    attribute_keys.append(key)
        sheet.append(["id", "name", *attribute_keys])
        for record, attributes in parsed:
            sheet.append(
                [record.ci_id, _safe_cell(record.name),
                 *[_safe_cell(attributes.get(key)) for key in attribute_keys]]
            )

    relations_sheet = workbook.create_sheet("关系")
    relations_sheet.append(["id", "type", "from_id", "to_id"])
    for record in iter_all_relations(session, dataset.id):
        relations_sheet.append([record.relation_id, record.type, record.from_ci_id, record.to_ci_id])

    quality_sheet = workbook.create_sheet("质量报告")
    quality_sheet.append(
        [
            "kind",
            "ci_type",
            "field",
            "requested_count",
            "affected_count",
            "ci_id",
            "source_id",
            "applied_value",
        ]
    )
    for row in _iter_quality_rows(json.loads(dataset.quality_report_json)):
        quality_sheet.append([_safe_cell(value) for value in row])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), f"dataset-{dataset.id}.xlsx"

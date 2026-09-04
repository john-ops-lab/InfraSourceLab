"""导出：JSON 与 CSV 保留 ID 与关系引用，且与 API 数据一致；XLSX 可生成。"""

import csv
import io
import json
import zipfile

import pytest
from openpyxl import load_workbook

from conftest import default_proposal


@pytest.fixture()
def dataset_id(client, auth):
    response = client.post(
        "/api/v1/datasets", headers=auth, json={"spec": default_proposal().spec}
    )
    assert response.status_code == 201
    return response.json()["id"]


def test_json_export_matches_api(client, auth, dataset_id):
    response = client.get(
        f"/api/v1/datasets/{dataset_id}/export", headers=auth, params={"format": "json"}
    )
    assert response.status_code == 200
    payload = json.loads(response.content)

    assert payload["dataset"]["record_count"] == 29
    assert len(payload["cis"]) == 29
    assert len(payload["relations"]) == 28

    # 与分页 API 的数据一致
    api_cis = client.get(
        f"/api/v1/datasets/{dataset_id}/cis", headers=auth, params={"page_size": 200}
    ).json()["items"]
    assert {item["id"] for item in api_cis} == {item["id"] for item in payload["cis"]}

    # 关系引用不悬空
    ci_ids = {item["id"] for item in payload["cis"]}
    for rel in payload["relations"]:
        assert rel["from_id"] in ci_ids
        assert rel["to_id"] in ci_ids

    # 不暴露内部 search_text
    assert "search_text" not in json.dumps(payload)


def test_csv_export_zip_structure(client, auth, dataset_id):
    response = client.get(
        f"/api/v1/datasets/{dataset_id}/export", headers=auth, params={"format": "csv"}
    )
    assert response.status_code == 200
    archive = zipfile.ZipFile(io.BytesIO(response.content))
    names = set(archive.namelist())
    assert "summary.csv" in names
    assert "relations.csv" in names
    assert "spec.json" in names
    assert "quality_report.csv" in names
    assert any(name.startswith("ci_") for name in names)

    relations_csv = archive.read("relations.csv").decode("utf-8")
    assert "rel-000001" in relations_csv

    vm_csv = archive.read("ci_virtual_machine.csv").decode("utf-8")
    assert "vm-0001" in vm_csv

    spec = json.loads(archive.read("spec.json"))
    assert spec["seed"] == 42


def test_exports_include_exact_quality_report(client, auth):
    spec = default_proposal().spec
    spec["quality_defects"] = [
        {"kind": "missing_field", "ci_type": "physical_server", "field": "status", "count": 8},
        {"kind": "wrong_value", "ci_type": "physical_server", "field": "status", "count": 2},
    ]
    created = client.post("/api/v1/datasets", headers=auth, json={"spec": spec}).json()
    dataset_id = created["id"]

    json_payload = client.get(
        f"/api/v1/datasets/{dataset_id}/export", headers=auth, params={"format": "json"}
    ).json()
    reports = json_payload["dataset"]["quality_report"]
    assert reports[0]["affected_count"] == 8
    assert len(reports[0]["affected_ids"]) == 8
    assert reports[1]["requested_count"] == 2
    assert reports[1]["affected_count"] == 0

    csv_response = client.get(
        f"/api/v1/datasets/{dataset_id}/export", headers=auth, params={"format": "csv"}
    )
    archive = zipfile.ZipFile(io.BytesIO(csv_response.content))
    report_rows = list(
        csv.DictReader(io.StringIO(archive.read("quality_report.csv").decode("utf-8")))
    )
    for ci_id in reports[0]["affected_ids"]:
        assert any(row["ci_id"] == ci_id for row in report_rows)
    zero_row = next(row for row in report_rows if row["kind"] == "wrong_value")
    assert zero_row["requested_count"] == "2"
    assert zero_row["affected_count"] == "0"
    assert zero_row["ci_id"] == ""

    xlsx_response = client.get(
        f"/api/v1/datasets/{dataset_id}/export", headers=auth, params={"format": "xlsx"}
    )
    workbook = load_workbook(io.BytesIO(xlsx_response.content), read_only=True)
    quality_rows = list(workbook["质量报告"].iter_rows(values_only=True))
    assert ("wrong_value", "physical_server", "status", "2", "0", None, None, "unknown") in quality_rows


def test_xlsx_export_generates(client, auth, dataset_id):
    response = client.get(
        f"/api/v1/datasets/{dataset_id}/export", headers=auth, params={"format": "xlsx"}
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    assert len(response.content) > 1000


def test_unknown_export_format_rejected(client, auth, dataset_id):
    response = client.get(
        f"/api/v1/datasets/{dataset_id}/export", headers=auth, params={"format": "parquet"}
    )
    assert response.status_code == 422


def test_formula_injection_escaped(client, auth):
    """名称以 = 开头时，CSV 单元格必须加单引号前缀。"""
    spec = default_proposal().spec
    spec["name"] = "=cmd|' /C calc'!A0"
    response = client.post("/api/v1/datasets", headers=auth, json={"spec": spec})
    assert response.status_code == 201
    dataset_id = response.json()["id"]

    archive = zipfile.ZipFile(
        io.BytesIO(
            client.get(
                f"/api/v1/datasets/{dataset_id}/export",
                headers=auth,
                params={"format": "csv"},
            ).content
        )
    )
    summary = archive.read("summary.csv").decode("utf-8")
    assert "'=cmd" in summary
